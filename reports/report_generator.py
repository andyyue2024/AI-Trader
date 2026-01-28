# -*- coding: utf-8 -*-
"""
报告生成器
支持 PDF 和 Excel 格式报告生成
"""

import io
import json
import logging
import os
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class ReportConfig:
    """报告配置"""
    title: str = "AI-Trader Performance Report"
    subtitle: str = ""
    author: str = "AI-Trader System"
    include_charts: bool = True
    include_trades: bool = True
    include_positions: bool = True
    include_risk_metrics: bool = True
    include_daily_stats: bool = True
    logo_path: Optional[str] = None
    template_path: Optional[str] = None
    output_dir: str = "./reports"


@dataclass
class ReportData:
    """报告数据"""
    # 基本信息
    report_date: date = field(default_factory=date.today)
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    symbols: List[str] = field(default_factory=list)

    # 账户概览
    initial_equity: float = 0.0
    final_equity: float = 0.0
    total_return: float = 0.0

    # 绩效指标
    sharpe_ratio: float = 0.0
    sortino_ratio: float = 0.0
    max_drawdown: float = 0.0
    volatility: float = 0.0

    # 交易统计
    total_trades: int = 0
    winning_trades: int = 0
    losing_trades: int = 0
    win_rate: float = 0.0
    profit_factor: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0

    # 执行指标
    fill_rate: float = 0.0
    avg_slippage: float = 0.0
    total_volume: float = 0.0
    avg_daily_volume: float = 0.0

    # 详细数据
    trades: List[Dict] = field(default_factory=list)
    positions: List[Dict] = field(default_factory=list)
    daily_stats: List[Dict] = field(default_factory=list)
    equity_curve: List[Dict] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "report_date": self.report_date.isoformat(),
            "period": {
                "start": self.period_start.isoformat() if self.period_start else None,
                "end": self.period_end.isoformat() if self.period_end else None
            },
            "symbols": self.symbols,
            "account": {
                "initial_equity": self.initial_equity,
                "final_equity": self.final_equity,
                "total_return": self.total_return
            },
            "performance": {
                "sharpe_ratio": self.sharpe_ratio,
                "sortino_ratio": self.sortino_ratio,
                "max_drawdown": self.max_drawdown,
                "volatility": self.volatility
            },
            "trading": {
                "total_trades": self.total_trades,
                "winning_trades": self.winning_trades,
                "losing_trades": self.losing_trades,
                "win_rate": self.win_rate,
                "profit_factor": self.profit_factor,
                "avg_win": self.avg_win,
                "avg_loss": self.avg_loss
            },
            "execution": {
                "fill_rate": self.fill_rate,
                "avg_slippage": self.avg_slippage,
                "total_volume": self.total_volume,
                "avg_daily_volume": self.avg_daily_volume
            }
        }


class BaseReportGenerator(ABC):
    """报告生成器基类"""

    def __init__(self, config: ReportConfig = None):
        self.config = config or ReportConfig()
        Path(self.config.output_dir).mkdir(parents=True, exist_ok=True)

    @abstractmethod
    def generate(self, data: ReportData) -> str:
        """生成报告，返回文件路径"""
        pass

    def _get_output_path(self, extension: str) -> str:
        """获取输出文件路径"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.{extension}"
        return str(Path(self.config.output_dir) / filename)


class PDFReportGenerator(BaseReportGenerator):
    """PDF 报告生成器"""

    def generate(self, data: ReportData) -> str:
        """生成 PDF 报告"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, mm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                Image, PageBreak
            )
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
        except ImportError:
            logger.error("reportlab not installed. Run: pip install reportlab")
            return self._generate_text_fallback(data)

        output_path = self._get_output_path("pdf")

        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10
        )

        elements = []

        # 标题
        elements.append(Paragraph(self.config.title, title_style))
        if self.config.subtitle:
            elements.append(Paragraph(self.config.subtitle, styles['Normal']))
        elements.append(Spacer(1, 20))

        # 报告日期
        elements.append(Paragraph(
            f"Report Date: {data.report_date.strftime('%Y-%m-%d')}",
            styles['Normal']
        ))
        if data.period_start and data.period_end:
            elements.append(Paragraph(
                f"Period: {data.period_start} to {data.period_end}",
                styles['Normal']
            ))
        elements.append(Spacer(1, 20))

        # 账户概览
        elements.append(Paragraph("Account Overview", heading_style))
        account_data = [
            ["Metric", "Value"],
            ["Initial Equity", f"${data.initial_equity:,.2f}"],
            ["Final Equity", f"${data.final_equity:,.2f}"],
            ["Total Return", f"{data.total_return:.2%}"],
            ["Trading Symbols", ", ".join(data.symbols) or "N/A"]
        ]
        account_table = Table(account_data, colWidths=[2.5*inch, 2.5*inch])
        account_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(account_table)
        elements.append(Spacer(1, 20))

        # 绩效指标
        if self.config.include_risk_metrics:
            elements.append(Paragraph("Performance Metrics", heading_style))
            perf_data = [
                ["Metric", "Value", "Target"],
                ["Sharpe Ratio", f"{data.sharpe_ratio:.2f}", "≥ 2.0"],
                ["Sortino Ratio", f"{data.sortino_ratio:.2f}", "-"],
                ["Max Drawdown", f"{data.max_drawdown:.2%}", "≤ 15%"],
                ["Volatility", f"{data.volatility:.2%}", "-"],
                ["Fill Rate", f"{data.fill_rate:.2%}", "≥ 95%"],
                ["Avg Slippage", f"{data.avg_slippage:.4%}", "≤ 0.2%"]
            ]
            perf_table = Table(perf_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
            perf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(perf_table)
            elements.append(Spacer(1, 20))

        # 交易统计
        elements.append(Paragraph("Trading Statistics", heading_style))
        trading_data = [
            ["Metric", "Value"],
            ["Total Trades", str(data.total_trades)],
            ["Winning Trades", str(data.winning_trades)],
            ["Losing Trades", str(data.losing_trades)],
            ["Win Rate", f"{data.win_rate:.2%}"],
            ["Profit Factor", f"{data.profit_factor:.2f}"],
            ["Average Win", f"${data.avg_win:.2f}"],
            ["Average Loss", f"${data.avg_loss:.2f}"],
            ["Total Volume", f"${data.total_volume:,.2f}"],
            ["Avg Daily Volume", f"${data.avg_daily_volume:,.2f}"]
        ]
        trading_table = Table(trading_data, colWidths=[2.5*inch, 2.5*inch])
        trading_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(trading_table)

        # 交易明细
        if self.config.include_trades and data.trades:
            elements.append(PageBreak())
            elements.append(Paragraph("Trade History", heading_style))

            trade_header = ["Time", "Symbol", "Side", "Qty", "Price", "P&L"]
            trade_rows = [trade_header]

            for trade in data.trades[:50]:  # 最多显示50笔
                trade_rows.append([
                    trade.get("timestamp", "")[:19],
                    trade.get("symbol", ""),
                    trade.get("side", ""),
                    str(trade.get("quantity", 0)),
                    f"${trade.get('price', 0):.2f}",
                    f"${trade.get('pnl', 0):.2f}"
                ])

            trade_table = Table(trade_rows, colWidths=[1.2*inch, 0.8*inch, 0.6*inch, 0.5*inch, 0.8*inch, 0.8*inch])
            trade_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            elements.append(trade_table)

        # 页脚
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            f"Generated by {self.config.author} on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Normal']
        ))

        doc.build(elements)
        logger.info(f"PDF report generated: {output_path}")
        return output_path

    def _generate_text_fallback(self, data: ReportData) -> str:
        """降级生成文本报告"""
        output_path = self._get_output_path("txt")

        content = f"""
{'='*60}
{self.config.title}
{'='*60}

Report Date: {data.report_date}
Period: {data.period_start} to {data.period_end}
Symbols: {', '.join(data.symbols)}

ACCOUNT OVERVIEW
----------------
Initial Equity: ${data.initial_equity:,.2f}
Final Equity: ${data.final_equity:,.2f}
Total Return: {data.total_return:.2%}

PERFORMANCE METRICS
-------------------
Sharpe Ratio: {data.sharpe_ratio:.2f} (Target: ≥2.0)
Sortino Ratio: {data.sortino_ratio:.2f}
Max Drawdown: {data.max_drawdown:.2%} (Target: ≤15%)
Volatility: {data.volatility:.2%}

TRADING STATISTICS
------------------
Total Trades: {data.total_trades}
Win Rate: {data.win_rate:.2%}
Profit Factor: {data.profit_factor:.2f}
Fill Rate: {data.fill_rate:.2%} (Target: ≥95%)
Avg Slippage: {data.avg_slippage:.4%} (Target: ≤0.2%)
Total Volume: ${data.total_volume:,.2f}
Avg Daily Volume: ${data.avg_daily_volume:,.2f}

{'='*60}
Generated by {self.config.author}
{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)

        logger.info(f"Text report generated: {output_path}")
        return output_path


class ExcelReportGenerator(BaseReportGenerator):
    """Excel 报告生成器"""

    def generate(self, data: ReportData) -> str:
        """生成 Excel 报告"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.chart import LineChart, Reference
        except ImportError:
            logger.error("openpyxl/pandas not installed. Run: pip install openpyxl pandas")
            return self._generate_csv_fallback(data)

        output_path = self._get_output_path("xlsx")

        wb = Workbook()

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 概览页
        ws_overview = wb.active
        ws_overview.title = "Overview"

        # 标题
        ws_overview['A1'] = self.config.title
        ws_overview['A1'].font = Font(bold=True, size=18)
        ws_overview.merge_cells('A1:D1')

        ws_overview['A2'] = f"Report Date: {data.report_date}"
        ws_overview['A3'] = f"Period: {data.period_start} to {data.period_end}"

        # 账户概览
        overview_data = [
            ["Metric", "Value", "Target", "Status"],
            ["Initial Equity", f"${data.initial_equity:,.2f}", "-", "-"],
            ["Final Equity", f"${data.final_equity:,.2f}", "-", "-"],
            ["Total Return", f"{data.total_return:.2%}", "-", "-"],
            ["Sharpe Ratio", f"{data.sharpe_ratio:.2f}", "≥ 2.0", "✓" if data.sharpe_ratio >= 2 else "✗"],
            ["Max Drawdown", f"{data.max_drawdown:.2%}", "≤ 15%", "✓" if data.max_drawdown <= 0.15 else "✗"],
            ["Fill Rate", f"{data.fill_rate:.2%}", "≥ 95%", "✓" if data.fill_rate >= 0.95 else "✗"],
            ["Avg Slippage", f"{data.avg_slippage:.4%}", "≤ 0.2%", "✓" if data.avg_slippage <= 0.002 else "✗"],
            ["Daily Volume", f"${data.avg_daily_volume:,.2f}", "≥ $50,000", "✓" if data.avg_daily_volume >= 50000 else "✗"],
        ]

        for row_idx, row in enumerate(overview_data, start=5):
            for col_idx, value in enumerate(row, start=1):
                cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align
                if row_idx == 5:
                    cell.font = header_font
                    cell.fill = header_fill

        # 调整列宽
        ws_overview.column_dimensions['A'].width = 20
        ws_overview.column_dimensions['B'].width = 20
        ws_overview.column_dimensions['C'].width = 15
        ws_overview.column_dimensions['D'].width = 10

        # 交易统计页
        ws_trading = wb.create_sheet("Trading Stats")
        trading_data = [
            ["Metric", "Value"],
            ["Total Trades", data.total_trades],
            ["Winning Trades", data.winning_trades],
            ["Losing Trades", data.losing_trades],
            ["Win Rate", f"{data.win_rate:.2%}"],
            ["Profit Factor", f"{data.profit_factor:.2f}"],
            ["Average Win", f"${data.avg_win:.2f}"],
            ["Average Loss", f"${data.avg_loss:.2f}"],
            ["Total Volume", f"${data.total_volume:,.2f}"],
        ]

        for row_idx, row in enumerate(trading_data, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws_trading.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        # 交易明细页
        if self.config.include_trades and data.trades:
            ws_trades = wb.create_sheet("Trades")

            headers = ["Timestamp", "Symbol", "Side", "Quantity", "Price", "P&L", "Commission"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_trades.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, trade in enumerate(data.trades, start=2):
                ws_trades.cell(row=row_idx, column=1, value=trade.get("timestamp", ""))
                ws_trades.cell(row=row_idx, column=2, value=trade.get("symbol", ""))
                ws_trades.cell(row=row_idx, column=3, value=trade.get("side", ""))
                ws_trades.cell(row=row_idx, column=4, value=trade.get("quantity", 0))
                ws_trades.cell(row=row_idx, column=5, value=trade.get("price", 0))
                ws_trades.cell(row=row_idx, column=6, value=trade.get("pnl", 0))
                ws_trades.cell(row=row_idx, column=7, value=trade.get("commission", 0))

        # 每日统计页
        if self.config.include_daily_stats and data.daily_stats:
            ws_daily = wb.create_sheet("Daily Stats")

            headers = ["Date", "Starting Equity", "Ending Equity", "Daily Return", "Trades", "Win Rate"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_daily.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row_idx, stat in enumerate(data.daily_stats, start=2):
                ws_daily.cell(row=row_idx, column=1, value=stat.get("date", ""))
                ws_daily.cell(row=row_idx, column=2, value=stat.get("starting_equity", 0))
                ws_daily.cell(row=row_idx, column=3, value=stat.get("ending_equity", 0))
                ws_daily.cell(row=row_idx, column=4, value=stat.get("daily_return", 0))
                ws_daily.cell(row=row_idx, column=5, value=stat.get("trade_count", 0))
                ws_daily.cell(row=row_idx, column=6, value=stat.get("win_rate", 0))

        wb.save(output_path)
        logger.info(f"Excel report generated: {output_path}")
        return output_path

    def _generate_csv_fallback(self, data: ReportData) -> str:
        """降级生成 CSV 报告"""
        import csv

        output_path = self._get_output_path("csv")

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            writer.writerow(["AI-Trader Performance Report"])
            writer.writerow([])
            writer.writerow(["Metric", "Value", "Target"])
            writer.writerow(["Initial Equity", f"${data.initial_equity:,.2f}", "-"])
            writer.writerow(["Final Equity", f"${data.final_equity:,.2f}", "-"])
            writer.writerow(["Total Return", f"{data.total_return:.2%}", "-"])
            writer.writerow(["Sharpe Ratio", f"{data.sharpe_ratio:.2f}", "≥ 2.0"])
            writer.writerow(["Max Drawdown", f"{data.max_drawdown:.2%}", "≤ 15%"])
            writer.writerow(["Fill Rate", f"{data.fill_rate:.2%}", "≥ 95%"])
            writer.writerow(["Avg Slippage", f"{data.avg_slippage:.4%}", "≤ 0.2%"])

        logger.info(f"CSV report generated: {output_path}")
        return output_path


class ReportScheduler:
    """报告调度器"""

    def __init__(self, generators: Dict[str, BaseReportGenerator] = None):
        self.generators = generators or {
            "pdf": PDFReportGenerator(),
            "excel": ExcelReportGenerator()
        }
        self._scheduled_tasks = []

    def generate_report(
        self,
        data: ReportData,
        format: str = "pdf"
    ) -> str:
        """生成报告"""
        generator = self.generators.get(format.lower())
        if not generator:
            raise ValueError(f"Unknown format: {format}. Available: {list(self.generators.keys())}")

        return generator.generate(data)

    def schedule_daily_report(
        self,
        data_provider,
        formats: List[str] = None,
        time_of_day: str = "18:00"
    ):
        """安排每日报告"""
        formats = formats or ["pdf", "excel"]

        async def task():
            import asyncio
            while True:
                # 计算到下一个执行时间的秒数
                now = datetime.now()
                target_hour, target_minute = map(int, time_of_day.split(":"))
                target_time = now.replace(hour=target_hour, minute=target_minute, second=0)

                if target_time <= now:
                    target_time += timedelta(days=1)

                wait_seconds = (target_time - now).total_seconds()
                await asyncio.sleep(wait_seconds)

                # 生成报告
                try:
                    data = data_provider()
                    for fmt in formats:
                        self.generate_report(data, fmt)
                except Exception as e:
                    logger.error(f"Failed to generate scheduled report: {e}")

        return task

    def send_report_email(
        self,
        report_path: str,
        recipients: List[str],
        smtp_config: Dict[str, Any]
    ) -> bool:
        """发送报告邮件"""
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.base import MIMEBase
            from email.mime.text import MIMEText
            from email import encoders

            msg = MIMEMultipart()
            msg['From'] = smtp_config.get('from', 'ai-trader@example.com')
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = f"AI-Trader Report - {date.today()}"

            body = "Please find the attached trading report."
            msg.attach(MIMEText(body, 'plain'))

            # 附件
            with open(report_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename={Path(report_path).name}'
            )
            msg.attach(part)

            # 发送
            with smtplib.SMTP(smtp_config['host'], smtp_config.get('port', 587)) as server:
                server.starttls()
                server.login(smtp_config['username'], smtp_config['password'])
                server.send_message(msg)

            logger.info(f"Report sent to {recipients}")
            return True

        except Exception as e:
            logger.error(f"Failed to send report email: {e}")
            return False


# 便捷函数
def generate_pdf_report(data: ReportData, config: ReportConfig = None) -> str:
    """生成 PDF 报告"""
    generator = PDFReportGenerator(config)
    return generator.generate(data)


def generate_excel_report(data: ReportData, config: ReportConfig = None) -> str:
    """生成 Excel 报告"""
    generator = ExcelReportGenerator(config)
    return generator.generate(data)
